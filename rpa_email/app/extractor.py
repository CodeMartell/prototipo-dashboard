"""
rpa_email/app/extractor.py
Extrai dados de planilhas KPI (.xlsx) e retorna estruturas
compativeis com o dashboard DataLens.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

LOGGER = logging.getLogger(__name__)

EXPECTED_MONTHS = {
    "Jan","Feb","Mar","Apr","May","Jun",
    "Jul","Aug","Sep","Oct","Nov","Dec",
}
EXPECTED_YEARS = {"Y24","Y25","Y26","Y27"}


# ---------------------------------------------------------------------------
# Modelos de saida
# ---------------------------------------------------------------------------

@dataclass
class KpiMonthlyRow:
    """Linha mensal para KPIs com target/result/achievement."""
    month: str
    year: str
    target: float
    result: float
    achievement: float
    kpi_key: str


@dataclass
class LogisticsVsProdRow:
    """Linha mensal para o KPI Logistics vs Production."""
    month: str
    year: str
    logistics_cost: float
    production_amount: float
    ratio: float


@dataclass
class ExtractionResult:
    logistic_cost: list[KpiMonthlyRow] = field(default_factory=list)
    air_freight: list[KpiMonthlyRow] = field(default_factory=list)
    logistics_vs_prod: list[LogisticsVsProdRow] = field(default_factory=list)
    incidental_cost: list[KpiMonthlyRow] = field(default_factory=list)
    total_cost: list[KpiMonthlyRow] = field(default_factory=list)
    demurrage: list[KpiMonthlyRow] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Utilitarios
# ---------------------------------------------------------------------------

def _safe_float(value: Any, label: str) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        LOGGER.warning("[EXTRACTION] Valor invalido em %s: %r", label, value)
        return None


def _read_sheet(path: Path) -> list[dict]:
    """Le a primeira aba de um .xlsx e retorna lista de dicts keyed pelo cabecalho."""
    LOGGER.info("[EXTRACTION] Lendo planilha: %s", path.name)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]


# ---------------------------------------------------------------------------
# Parsers por KPI
# ---------------------------------------------------------------------------

def _parse_kpi_row(raw: dict, kpi_key: str) -> KpiMonthlyRow | None:
    month = str(raw.get("month", "")).strip()
    year  = str(raw.get("year",  "")).strip()
    if month not in EXPECTED_MONTHS or year not in EXPECTED_YEARS:
        LOGGER.debug("[EXTRACTION] Linha ignorada — month=%s year=%s", month, year)
        return None
    target      = _safe_float(raw.get("target"),      f"{kpi_key}.target")
    result      = _safe_float(raw.get("result"),      f"{kpi_key}.result")
    achievement = _safe_float(raw.get("achievement"), f"{kpi_key}.achievement")
    if result is None or target is None:
        return None
    if achievement is None and result > 0:
        achievement = round(target / result, 4)
    return KpiMonthlyRow(
        month=month, year=year,
        target=round(target, 6), result=round(result, 6),
        achievement=round(achievement or 0, 4),
        kpi_key=kpi_key,
    )


def _parse_logistics_vs_prod_row(raw: dict) -> LogisticsVsProdRow | None:
    month = str(raw.get("month", "")).strip()
    year  = str(raw.get("year",  "")).strip()
    if month not in EXPECTED_MONTHS or year not in EXPECTED_YEARS:
        return None
    logistics_cost     = _safe_float(raw.get("logisticsCost"),     "logisticsCost")
    production_amount  = _safe_float(raw.get("productionAmount"),  "productionAmount")
    ratio              = _safe_float(raw.get("ratio"),             "ratio")
    if logistics_cost is None or production_amount is None:
        return None
    if ratio is None and production_amount > 0:
        ratio = round(logistics_cost / production_amount, 6)
    return LogisticsVsProdRow(
        month=month, year=year,
        logistics_cost=round(logistics_cost, 4),
        production_amount=round(production_amount, 4),
        ratio=round(ratio or 0, 6),
    )


def _read_sheet_by_name(path: Path, sheet_name: str | None = None) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]


class KpiExtractor:
    """
    Recebe uma pasta contendo as planilhas KPI e retorna ExtractionResult.
    Suporta arquivos individuais, arquivos consolidados com abas e identificação por nome/coluna.
    """

    _STANDARD_KPIS = ["logistic_cost", "air_freight", "incidental_cost", "total_cost", "demurrage"]

    def __init__(self, reports_dir: Path):
        self._dir = reports_dir

    def extract(self) -> ExtractionResult:
        result = ExtractionResult()
        LOGGER.info("[EXTRACTION] Iniciando extracao de KPIs em: %s", self._dir)

        xlsx_files = sorted(set(list(self._dir.glob("*.xlsx")) + list(self._dir.glob("*.XLSX"))))
        if not xlsx_files:
            LOGGER.warning("[EXTRACTION] Nenhum arquivo .xlsx encontrado em %s", self._dir)
            return result

        for path in xlsx_files:
            try:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                sheet_names = wb.sheetnames
                wb.close()

                LOGGER.info("[EXTRACTION] Analisando arquivo: %s (abas: %s)", path.name, sheet_names)

                for s_name in sheet_names:
                    raw_rows = _read_sheet_by_name(path, s_name)
                    if not raw_rows:
                        continue

                    # Determinar o KPI correspondente baseado no nome da aba, do arquivo ou das colunas
                    target_kpi = None
                    file_or_sheet = f"{path.stem.lower()} {s_name.lower()}"

                    if "logistics_vs_prod" in file_or_sheet or "prod" in file_or_sheet or "production" in file_or_sheet:
                        target_kpi = "logistics_vs_prod"
                    elif "logistic" in file_or_sheet or "war_room" in file_or_sheet or "war room" in file_or_sheet:
                        target_kpi = "logistic_cost"
                    elif "air" in file_or_sheet or "freight" in file_or_sheet:
                        target_kpi = "air_freight"
                    elif "incidental" in file_or_sheet:
                        target_kpi = "incidental_cost"
                    elif "total" in file_or_sheet:
                        target_kpi = "total_cost"
                    elif "demurrage" in file_or_sheet:
                        target_kpi = "demurrage"

                    # Se ainda nao detectou, inspeciona as colunas
                    if not target_kpi and raw_rows:
                        first_keys = set(raw_rows[0].keys())
                        if "productionAmount" in first_keys or "logisticsCost" in first_keys:
                            target_kpi = "logistics_vs_prod"
                        elif "target" in first_keys and "result" in first_keys:
                            target_kpi = "logistic_cost"

                    if not target_kpi:
                        continue

                    parsed_count = 0
                    if target_kpi == "logistics_vs_prod":
                        for raw in raw_rows:
                            row = _parse_logistics_vs_prod_row(raw)
                            if row:
                                result.logistics_vs_prod.append(row)
                                parsed_count += 1
                    else:
                        for raw in raw_rows:
                            row = _parse_kpi_row(raw, target_kpi)
                            if row:
                                getattr(result, target_kpi).append(row)
                                parsed_count += 1

                    LOGGER.info(
                        "[EXTRACTION] %s [%s] -> %s: %d registros extraidos",
                        path.name, s_name, target_kpi, parsed_count,
                    )
            except Exception as exc:
                msg = f"Erro ao processar {path.name}: {exc}"
                LOGGER.exception("[EXTRACTION] %s", msg)
                result.errors.append(msg)

        LOGGER.info(
            "[EXTRACTION] Extracao concluida. logistic_cost=%d air_freight=%d "
            "logistics_vs_prod=%d incidental_cost=%d total_cost=%d demurrage=%d erros=%d",
            len(result.logistic_cost), len(result.air_freight),
            len(result.logistics_vs_prod), len(result.incidental_cost),
            len(result.total_cost), len(result.demurrage), len(result.errors),
        )
        return result

