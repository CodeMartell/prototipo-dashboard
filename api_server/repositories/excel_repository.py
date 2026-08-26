from __future__ import annotations

import logging
from pathlib import Path
from typing import Any
import openpyxl

from .base import KpiRepositoryInterface

LOGGER = logging.getLogger(__name__)


class ExcelKpiRepository(KpiRepositoryInterface):
    """Implementação baseada em cache local Excel (dados_dashboard.xlsx)."""

    def __init__(self, excel_path: Path):
        self._path = excel_path

    def is_available(self) -> bool:
        return self._path.exists()

    def fetch_all(self) -> dict[str, list[dict[str, Any]]]:
        if not self._path.exists():
            return {}

        data: dict[str, list[dict[str, Any]]] = {}
        standard_sheets = ["logistic_cost", "air_freight", "incidental_cost", "total_cost", "demurrage"]

        try:
            wb = openpyxl.load_workbook(self._path, read_only=True, data_only=True)
            for sheet in standard_sheets:
                if sheet in wb.sheetnames:
                    ws = wb[sheet]
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) > 1:
                        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
                        items = []
                        for row in rows[1:]:
                            d = dict(zip(headers, row))
                            if d.get("month") and d.get("year"):
                                try:
                                    items.append({
                                        "month": str(d["month"]).strip(),
                                        "year": str(d["year"]).strip(),
                                        "target": float(d["target"]) if d.get("target") is not None else 0.0,
                                        "result": float(d["result"]) if d.get("result") is not None else 0.0,
                                        "achievement": float(d["achievement"]) if d.get("achievement") is not None else None,
                                    })
                                except (ValueError, TypeError):
                                    pass
                        data[sheet] = items

            if "logistics_vs_prod" in wb.sheetnames:
                ws = wb["logistics_vs_prod"]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
                    items = []
                    for row in rows[1:]:
                        d = dict(zip(headers, row))
                        if d.get("month") and d.get("year"):
                            try:
                                items.append({
                                    "month": str(d["month"]).strip(),
                                    "year": str(d["year"]).strip(),
                                    "logisticsCost": float(d["logisticsCost"]) if d.get("logisticsCost") is not None else 0.0,
                                    "productionAmount": float(d["productionAmount"]) if d.get("productionAmount") is not None else 0.0,
                                    "ratio": float(d["ratio"]) if d.get("ratio") is not None else None,
                                })
                            except (ValueError, TypeError):
                                pass
                    data["logistics_vs_prod"] = items

            wb.close()
        except Exception as exc:
            LOGGER.error("Erro ao ler dados do arquivo Excel %s: %s", self._path, exc)

        return data
