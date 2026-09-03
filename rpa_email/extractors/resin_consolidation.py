from pathlib import Path

from openpyxl import load_workbook

from rpa_email.extractors.common import MONTHS, monthly_record, parse_number


KPI_KEY = "incidental_cost"
DISPLAY_NAME = "Resin Consolidation"


def extract(source: Path) -> list[dict]:
    """Extrai o saving de Resin Consolidation do arquivo 3-indicadores."""

    workbook = load_workbook(source, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        records = []
        for year, first_column in (("Y25", 6), ("Y26", 20)):
            for index, month in enumerate(MONTHS):
                column = first_column + index
                target = sheet.cell(23, column).value
                cost = sheet.cell(24, column).value
                tax = sheet.cell(25, column).value
                if target is None or cost is None or tax is None:
                    continue
                target_value = parse_number(target)
                result = round(target_value - parse_number(cost) - parse_number(tax), 2)
                records.append(monthly_record(month, year, target_value, result, lower_is_better=False))
        return records
    finally:
        workbook.close()
