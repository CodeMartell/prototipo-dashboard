"""
api_server/server.py
Servidor Flask local que serve os dados KPI do PostgreSQL ao dashboard.

Rotas:
    GET /api/dashboard  — retorna todos os KPIs como JSON
    GET /api/health     — status do servidor

Uso:
    python api_server/server.py
"""
from __future__ import annotations

import logging
import os
from pathlib import Path

from dotenv import load_dotenv
from flask import Flask, jsonify
from flask_cors import CORS

# ---------------------------------------------------------------------------
# Ambiente
# ---------------------------------------------------------------------------
ROOT_DIR = Path(__file__).resolve().parents[1]
load_dotenv(ROOT_DIR / ".env")
load_dotenv(ROOT_DIR / "rpa_email" / ".env")

DATABASE_URL = os.getenv("DATABASE_URL", "").strip()

# ---------------------------------------------------------------------------
# App Flask
# ---------------------------------------------------------------------------
app = Flask(__name__)
CORS(app, origins=["http://localhost:5173", "http://127.0.0.1:5173"])

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)
LOGGER = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_kpi_data_from_excel(path: Path) -> dict:
    """Le todos os KPIs a partir do arquivo Excel dados_dashboard.xlsx."""
    import openpyxl
    if not path.exists():
        return {}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    data: dict = {}
    standard_sheets = ["logistic_cost", "air_freight", "incidental_cost", "total_cost", "demurrage"]

    for sheet in standard_sheets:
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) > 1:
                headers = [str(h).strip() if h is not None else "" for h in rows[0]]
                items = []
                for row in rows[1:]:
                    d = dict(zip(headers, row))
                    if d.get("month") and d.get("year"):
                        try:
                            items.append({
                                "month": str(d["month"]).strip(),
                                "year": str(d["year"]).strip(),
                                "target": float(d["target"]) if d.get("target") is not None else 0.0,
                                "result": float(d["result"]) if d.get("result") is not None else 0.0,
                                "achievement": float(d["achievement"]) if d.get("achievement") is not None else None,
                            })
                        except (ValueError, TypeError):
                            pass
                data[sheet] = items

    if "logistics_vs_prod" in wb.sheetnames:
        ws = wb["logistics_vs_prod"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            items = []
            for row in rows[1:]:
                d = dict(zip(headers, row))
                if d.get("month") and d.get("year"):
                    try:
                        items.append({
                            "month": str(d["month"]).strip(),
                            "year": str(d["year"]).strip(),
                            "logisticsCost": float(d["logisticsCost"]) if d.get("logisticsCost") is not None else 0.0,
                            "productionAmount": float(d["productionAmount"]) if d.get("productionAmount") is not None else 0.0,
                            "ratio": float(d["ratio"]) if d.get("ratio") is not None else None,
                        })
                    except (ValueError, TypeError):
                        pass
            data["logistics_vs_prod"] = items

    wb.close()
    return data


def _get_kpi_data() -> tuple[dict, str]:
    """Le todos os KPIs: tenta PostgreSQL primeiro; fallback para Excel / kpi_reports."""
    # 1. Tenta PostgreSQL se DATABASE_URL estiver configurada
    if DATABASE_URL:
        try:
            import psycopg
            data: dict = {}
            STANDARD_TABLES = {
                "logistic_cost":   "kpi_logistic_cost",
                "air_freight":     "kpi_air_freight",
                "incidental_cost": "kpi_incidental_cost",
                "total_cost":      "kpi_total_cost",
                "demurrage":       "kpi_demurrage",
            }
            # Conecta com timeout curto de 2 segundos
            with psycopg.connect(DATABASE_URL, connect_timeout=2) as conn:
                for kpi_key, table in STANDARD_TABLES.items():
                    rows = conn.execute(
                        f"SELECT month, year, target, result, achievement "
                        f"FROM {table} ORDER BY year, month"
                    ).fetchall()
                    data[kpi_key] = [
                        {
                            "month": r[0], "year": r[1],
                            "target": float(r[2]), "result": float(r[3]),
                            "achievement": float(r[4]) if r[4] is not None else None,
                        }
                        for r in rows
                    ]

                rows = conn.execute(
                    "SELECT month, year, logistics_cost, production_amount, ratio "
                    "FROM kpi_logistics_vs_prod ORDER BY year, month"
                ).fetchall()
                data["logistics_vs_prod"] = [
                    {
                        "month": r[0], "year": r[1],
                        "logisticsCost": float(r[2]),
                        "productionAmount": float(r[3]),
                        "ratio": float(r[4]) if r[4] is not None else None,
                    }
                    for r in rows
                ]
            total = sum(len(v) for v in data.values())
            if total > 0:
                return data, "postgresql"
        except Exception as exc:
            LOGGER.warning("[API] PostgreSQL inacessível (%s), tentando cache local Excel...", exc)

    # 2. Tenta ler do arquivo dados_dashboard.xlsx
    excel_path = ROOT_DIR / "dados_dashboard.xlsx"
    if excel_path.exists():
        excel_data = _get_kpi_data_from_excel(excel_path)
        total = sum(len(v) for v in excel_data.values())
        if total > 0:
            return excel_data, "excel_cache"

    # 3. Fallback: extrai diretamente da pasta kpi_reports/
    kpi_dir = ROOT_DIR / "kpi_reports"
    if kpi_dir.exists():
        import sys
        sys.path.insert(0, str(ROOT_DIR))
        from rpa_email.app.extractor import KpiExtractor
        extractor = KpiExtractor(kpi_dir)
        ext = extractor.extract()
        data = {
            "logistic_cost": [
                {"month": r.month, "year": r.year, "target": r.target, "result": r.result, "achievement": r.achievement}
                for r in ext.logistic_cost
            ],
            "air_freight": [
                {"month": r.month, "year": r.year, "target": r.target, "result": r.result, "achievement": r.achievement}
                for r in ext.air_freight
            ],
            "incidental_cost": [
                {"month": r.month, "year": r.year, "target": r.target, "result": r.result, "achievement": r.achievement}
                for r in ext.incidental_cost
            ],
            "total_cost": [
                {"month": r.month, "year": r.year, "target": r.target, "result": r.result, "achievement": r.achievement}
                for r in ext.total_cost
            ],
            "demurrage": [
                {"month": r.month, "year": r.year, "target": r.target, "result": r.result, "achievement": r.achievement}
                for r in ext.demurrage
            ],
            "logistics_vs_prod": [
                {"month": r.month, "year": r.year, "logisticsCost": r.logistics_cost, "productionAmount": r.production_amount, "ratio": r.ratio}
                for r in ext.logistics_vs_prod
            ],
        }
        return data, "kpi_reports"

    return {}, "none"


# ---------------------------------------------------------------------------
# Rotas
# ---------------------------------------------------------------------------

@app.get("/api/dashboard")
def dashboard():
    """Retorna todos os dados KPI para o frontend."""
    try:
        data, source = _get_kpi_data()
        total = sum(len(v) for v in data.values())
        LOGGER.info("[API] GET /api/dashboard — %d registros retornados (fonte: %s).", total, source)
        return jsonify(data)
    except Exception as exc:
        LOGGER.exception("[API] Erro ao obter dados: %s", exc)
        return jsonify({"error": str(exc)}), 500


@app.get("/api/health")
def health():
    """Verifica se o servidor esta ativo e se o banco ou planilha esta acessivel."""
    _, source = _get_kpi_data()
    status = {"status": "ok", "source": source}
    LOGGER.info("[API] GET /api/health — ok (fonte: %s)", source)
    return jsonify(status), 200


# ---------------------------------------------------------------------------
# Ponto de entrada
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    port = int(os.getenv("API_PORT", "5001"))
    LOGGER.info("[API] Iniciando servidor na porta %d...", port)
    app.run(host="0.0.0.0", port=port, debug=False)

