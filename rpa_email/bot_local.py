"""
rpa_email/bot_local.py
Ponto de entrada para o fluxo local de teste:

  Planilhas KPI → Extracao → PostgreSQL → dados_dashboard.xlsx

NAO requer servidor IMAP real. Simula o recebimento de e-mail
com anexos processando diretamente as planilhas da pasta kpi_reports/.

Uso:
    cd <raiz_do_projeto>
    python -m rpa_email.bot_local
"""
from __future__ import annotations

import logging
import os
from pathlib import Path

from dotenv import load_dotenv

from rpa_email.app.extractor import KpiExtractor
from rpa_email.app.kpi_repository import KpiPostgresRepository

# ---------------------------------------------------------------------------
# Configuracao de logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)
LOGGER = logging.getLogger(__name__)

ROOT_DIR = Path(__file__).resolve().parent
load_dotenv(ROOT_DIR.parent / ".env")
load_dotenv(ROOT_DIR / ".env")

DATABASE_URL     = os.getenv("DATABASE_URL", "").strip()
KPI_REPORTS_DIR  = Path(os.getenv("KPI_REPORTS_DIR", str(ROOT_DIR.parent / "kpi_reports")))
EXCEL_OUTPUT_PATH = Path(os.getenv("EXCEL_OUTPUT_PATH", str(ROOT_DIR.parent / "dados_dashboard.xlsx")))

# ---------------------------------------------------------------------------
# Escrita no Excel de cache local
# ---------------------------------------------------------------------------

def _write_excel_cache(data: dict, path: Path) -> None:
    """Grava cache local Excel com todos os KPIs para inspecao manual."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    LOGGER.info("[EXCEL] Atualizando cache local: %s", path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HEADER_FILL = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    DATA_FONT   = Font(name="Calibri", size=10)
    CENTER      = Alignment(horizontal="center", vertical="center")

    sheet_configs = {
        "logistic_cost":    ("logistic_cost",    ["month","year","target","result","achievement"]),
        "air_freight":      ("air_freight",       ["month","year","target","result","achievement"]),
        "logistics_vs_prod":("logistics_vs_prod", ["month","year","logisticsCost","productionAmount","ratio"]),
        "incidental_cost":  ("incidental_cost",   ["month","year","target","result","achievement"]),
        "total_cost":       ("total_cost",        ["month","year","target","result","achievement"]),
        "demurrage":        ("demurrage",         ["month","year","target","result","achievement"]),
    }

    for key, (sheet_name, cols) in sheet_configs.items():
        raw_rows = data.get(key, [])
        # Deduplica por (month, year)
        dedup_map = {}
        for r in raw_rows:
            m, y = r.get("month"), r.get("year")
            if m and y:
                dedup_map[(str(m).strip(), str(y).strip())] = r
        rows = list(dedup_map.values())

        ws = wb.create_sheet(title=sheet_name)
        ws.freeze_panes = "A2"
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = CENTER
            ws.column_dimensions[get_column_letter(ci)].width = max(len(col)+6, 14)
        for ri, row in enumerate(rows, 2):
            for ci, col in enumerate(cols, 1):
                cell = ws.cell(row=ri, column=ci, value=row.get(col))
                cell.font = DATA_FONT; cell.alignment = CENTER

    wb.save(path)
    LOGGER.info("[EXCEL] Cache local salvo com sucesso: %s", path)


# ---------------------------------------------------------------------------
# Fluxo principal
# ---------------------------------------------------------------------------

def main() -> int:
    LOGGER.info("=" * 60)
    LOGGER.info("[BOT] DataLens — Bot Local (modo sem IMAP)")
    LOGGER.info("=" * 60)

    # Validar DATABASE_URL
    if not DATABASE_URL:
        LOGGER.error("[BOT] Variavel DATABASE_URL nao configurada. Configure o .env e tente novamente.")
        return 1

    # Verificar diretorio de planilhas
    if not KPI_REPORTS_DIR.exists():
        LOGGER.error(
            "[BOT] Diretorio kpi_reports nao encontrado: %s\n"
            "      Execute primeiro: python scripts/generate_kpi_reports.py",
            KPI_REPORTS_DIR,
        )
        return 1

    LOGGER.info("[EMAIL] Simulando recebimento de e-mail com anexos KPI...")
    LOGGER.info("[EMAIL] Diretorio de planilhas: %s", KPI_REPORTS_DIR)

    # 1. Extracao
    LOGGER.info("[EXTRACTION] Iniciando extracao das planilhas...")
    extractor = KpiExtractor(KPI_REPORTS_DIR)
    result = extractor.extract()

    if result.errors:
        for err in result.errors:
            LOGGER.warning("[EXTRACTION] Aviso: %s", err)

    # Verificar se extraiu algo util
    total_extracted = sum([
        len(result.logistic_cost), len(result.air_freight),
        len(result.logistics_vs_prod), len(result.incidental_cost),
        len(result.total_cost), len(result.demurrage),
    ])
    if total_extracted == 0:
        LOGGER.error("[EXTRACTION] Nenhum dado extraido. Encerrando.")
        return 1

    LOGGER.info("[VALIDATION] Total de registros extraidos: %d", total_extracted)
    LOGGER.info("[VALIDATION] Dados validos.")

    # 2. Persistencia no PostgreSQL
    LOGGER.info("[DB] Conectando ao PostgreSQL...")
    kpi_repo = KpiPostgresRepository(DATABASE_URL)

    try:
        kpi_repo.initialize()
    except Exception as exc:
        LOGGER.error("[DB] Falha ao inicializar schema: %s", exc)
        return 1

    kpi_map = {
        "logistic_cost":   result.logistic_cost,
        "air_freight":     result.air_freight,
        "incidental_cost": result.incidental_cost,
        "total_cost":      result.total_cost,
        "demurrage":       result.demurrage,
    }

    errors = 0
    for kpi_key, rows in kpi_map.items():
        try:
            kpi_repo.upsert_standard_kpi(kpi_key, rows)
        except Exception as exc:
            LOGGER.error("[DB] Falha ao gravar %s: %s", kpi_key, exc)
            errors += 1

    try:
        kpi_repo.upsert_logistics_vs_prod(result.logistics_vs_prod)
    except Exception as exc:
        LOGGER.error("[DB] Falha ao gravar logistics_vs_prod: %s", exc)
        errors += 1

    if errors:
        LOGGER.warning("[DB] %d KPI(s) com erro de gravacao.", errors)
    else:
        LOGGER.info("[DB] Todos os KPIs gravados com sucesso no PostgreSQL.")

    # 3. Cache local Excel
    try:
        all_data = kpi_repo.fetch_all()
        _write_excel_cache(all_data, EXCEL_OUTPUT_PATH)
        LOGGER.info("[EXCEL] dados_dashboard.xlsx atualizado.")
    except Exception as exc:
        LOGGER.warning("[EXCEL] Nao foi possivel atualizar o cache local: %s", exc)

    LOGGER.info("[API] Dados disponiveis para o dashboard via GET /api/dashboard")
    LOGGER.info("=" * 60)
    LOGGER.info("[BOT] Execucao concluida. Erros: %d", errors)
    LOGGER.info("=" * 60)
    return 1 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
