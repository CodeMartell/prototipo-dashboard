"""Extrai planilhas já normalizadas para o contrato do dashboard."""
from __future__ import annotations

import logging
import math
from pathlib import Path
from typing import Any

import openpyxl

from rpa_email.extractors.logistics_cost_vs_prod_amount import parse_normalized
from rpa_email.extractors.models import ExtractionResult, KpiMonthlyRow

LOGGER = logging.getLogger(__name__)

EXPECTED_MONTHS = {
    "Jan","Feb","Mar","Apr","May","Jun",
    "Jul","Aug","Sep","Oct","Nov","Dec",
}
EXPECTED_YEARS = {"Y24","Y25","Y26","Y27"}
LOWER_IS_BETTER = {
    "logistic_cost": True,
    "air_freight": True,
    "incidental_cost": False,
    "total_cost": False,
    "demurrage": True,
}


# ---------------------------------------------------------------------------
# Utilitarios
# ---------------------------------------------------------------------------

def _safe_float(value: Any, label: str) -> float | None:
    try:
        number = float(value)
        if not math.isfinite(number) or number < 0:
            raise ValueError
        return number
    except (TypeError, ValueError):
        LOGGER.warning("[EXTRACTION] Valor invalido em %s: %r", label, value)
        return None


# ---------------------------------------------------------------------------
# Parsers por KPI
# ---------------------------------------------------------------------------

def _parse_kpi_row(raw: dict, kpi_key: str) -> KpiMonthlyRow | None:
    month = str(raw.get("month", "")).strip()
    year  = str(raw.get("year",  "")).strip()
    if month not in EXPECTED_MONTHS or year not in EXPECTED_YEARS:
        LOGGER.debug("[EXTRACTION] Linha ignorada — month=%s year=%s", month, year)
        return None
    target      = _safe_float(raw.get("target"),      f"{kpi_key}.target")
    result      = _safe_float(raw.get("result"),      f"{kpi_key}.result")
    raw_achievement = raw.get("achievement")
    achievement = (
        _safe_float(raw_achievement, f"{kpi_key}.achievement")
        if raw_achievement not in (None, "")
        else None
    )
    if result is None or target is None:
        return None
    if achievement is None:
        lower_is_better = LOWER_IS_BETTER[kpi_key]
        numerator, denominator = (target, result) if lower_is_better else (result, target)
        achievement = 1.0 if numerator == denominator == 0 else round(numerator / denominator, 4) if denominator > 0 else None
    return KpiMonthlyRow(
        month=month, year=year,
        target=round(target, 6), result=round(result, 6),
        achievement=round(achievement, 4) if achievement is not None else None,
        kpi_key=kpi_key,
    )


def _read_sheet_by_name(path: Path, sheet_name: str | None = None) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]


class KpiExtractor:
    """
    Recebe uma pasta contendo as planilhas KPI e retorna ExtractionResult.
    Suporta arquivos individuais, arquivos consolidados com abas e identificação por nome/coluna.
    """

    _STANDARD_KPIS = ["logistic_cost", "air_freight", "incidental_cost", "total_cost", "demurrage"]

    def __init__(self, reports_dir: Path):
        self._dir = reports_dir

    def extract(self) -> ExtractionResult:
        result = ExtractionResult()
        LOGGER.info("[EXTRACTION] Iniciando extracao de KPIs em: %s", self._dir)

        xlsx_files = sorted(set(list(self._dir.glob("*.xlsx")) + list(self._dir.glob("*.XLSX"))))
        if not xlsx_files:
            LOGGER.warning("[EXTRACTION] Nenhum arquivo .xlsx encontrado em %s", self._dir)
            return result

        for path in xlsx_files:
            try:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                sheet_names = wb.sheetnames
                wb.close()

                LOGGER.info("[EXTRACTION] Analisando arquivo: %s (abas: %s)", path.name, sheet_names)

                for s_name in sheet_names:
                    raw_rows = _read_sheet_by_name(path, s_name)
                    if not raw_rows:
                        result.errors.append(f"{path.name} [{s_name}]: planilha sem registros")
                        continue

                    # Determinar o KPI correspondente baseado no nome da aba, do arquivo ou das colunas
                    target_kpi = None
                    file_or_sheet = f"{path.stem.lower()} {s_name.lower()}"

                    if "logistics_vs_prod" in file_or_sheet or "prod" in file_or_sheet or "production" in file_or_sheet:
                        target_kpi = "logistics_vs_prod"
                    elif "logistic" in file_or_sheet or "war_room" in file_or_sheet or "war room" in file_or_sheet:
                        target_kpi = "logistic_cost"
                    elif "air" in file_or_sheet or "freight" in file_or_sheet:
                        target_kpi = "air_freight"
                    elif "incidental" in file_or_sheet:
                        target_kpi = "incidental_cost"
                    elif "total" in file_or_sheet:
                        target_kpi = "total_cost"
                    elif "demurrage" in file_or_sheet:
                        target_kpi = "demurrage"

                    # Se ainda nao detectou, inspeciona as colunas
                    if not target_kpi and raw_rows:
                        first_keys = set(raw_rows[0].keys())
                        if "productionAmount" in first_keys or "logisticsCost" in first_keys:
                            target_kpi = "logistics_vs_prod"
                        elif "target" in first_keys and "result" in first_keys:
                            target_kpi = "logistic_cost"

                    if not target_kpi:
                        result.errors.append(f"{path.name} [{s_name}]: tipo de KPI não identificado")
                        continue

                    required = ({"month", "year", "logisticsCost", "productionAmount"}
                                if target_kpi == "logistics_vs_prod"
                                else {"month", "year", "target", "result"})
                    headers = set(raw_rows[0])
                    missing = sorted(required - headers)
                    if missing:
                        result.errors.append(
                            f"{path.name} [{s_name}]: colunas obrigatórias ausentes: {', '.join(missing)}"
                        )
                        continue

                    parsed_count = 0
                    if target_kpi == "logistics_vs_prod":
                        for row_number, raw in enumerate(raw_rows, start=2):
                            row = parse_normalized(raw, EXPECTED_MONTHS, EXPECTED_YEARS)
                            if row:
                                result.logistics_vs_prod.append(row)
                                parsed_count += 1
                            else:
                                result.errors.append(f"{path.name} [{s_name}] linha {row_number}: dados inválidos")
                    else:
                        for row_number, raw in enumerate(raw_rows, start=2):
                            row = _parse_kpi_row(raw, target_kpi)
                            if row:
                                getattr(result, target_kpi).append(row)
                                parsed_count += 1
                            else:
                                result.errors.append(f"{path.name} [{s_name}] linha {row_number}: dados inválidos")

                    LOGGER.info(
                        "[EXTRACTION] %s [%s] -> %s: %d registros extraidos",
                        path.name, s_name, target_kpi, parsed_count,
                    )
            except Exception as exc:
                msg = f"Erro ao processar {path.name}: {exc}"
                LOGGER.exception("[EXTRACTION] %s", msg)
                result.errors.append(msg)

        if result.errors:
            # O e-mail é uma unidade atômica: nunca enviar só as linhas válidas.
            for name in (*self._STANDARD_KPIS, "logistics_vs_prod"):
                getattr(result, name).clear()

        LOGGER.info(
            "[EXTRACTION] Extracao concluida. logistic_cost=%d air_freight=%d "
            "logistics_vs_prod=%d incidental_cost=%d total_cost=%d demurrage=%d erros=%d",
            len(result.logistic_cost), len(result.air_freight),
            len(result.logistics_vs_prod), len(result.incidental_cost),
            len(result.total_cost), len(result.demurrage), len(result.errors),
        )
        return result

