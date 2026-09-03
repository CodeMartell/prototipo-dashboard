"""Extrai os KPIs solicitados e gera as seis planilhas oficiais em saidas_kpi.

Dependências:
    pip install pyxlsb openpyxl

Uso:
    python extrair_kpis.py
    python extrair_kpis.py --origem "C:\\...\\dados-reais" --saida "C:\\...\\saidas_kpi"
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

# Permite usar a cópia local instalada junto com este projeto.
LOCAL_DEPS = Path(__file__).resolve().parent / ".python_deps"
if LOCAL_DEPS.exists():
    sys.path.insert(0, str(LOCAL_DEPS))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from pyxlsb import open_workbook


DEFAULT_SOURCE = Path(r"C:\Users\ROMULO_LIRA\Desktop\dados-reais")
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "saidas_kpi"
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

NAVY = "17365D"
BLUE = "4472C4"
LIGHT_BLUE = "D9EAF7"
GRAY = "A6A6A6"
LIGHT_GRAY = "E7E6E6"
RED = "C00000"
LIGHT_RED = "FCE4D6"
WHITE = "FFFFFF"
GREEN = "70AD47"
THIN_GRAY = Side(style="thin", color="D9E1F2")


def find_file(folder: Path, fragment: str) -> Path:
    matches = [p for p in folder.glob("*.xlsb") if not p.name.startswith("~$") and fragment.lower() in p.name.lower()]
    if not matches:
        raise FileNotFoundError(f"Arquivo contendo {fragment!r} não encontrado em {folder}")
    return max(matches, key=lambda p: p.stat().st_mtime)


def find_xlsx(folder: Path, fragment: str) -> Path:
    matches = [p for p in folder.glob("*.xlsx") if not p.name.startswith("~$") and fragment.lower() in p.name.lower()]
    if not matches:
        raise FileNotFoundError(f"Arquivo XLSX contendo {fragment!r} não encontrado em {folder}")
    return max(matches, key=lambda p: p.stat().st_mtime)


def find_any_file(folder: Path, fragment: str) -> Path:
    """Localiza o arquivo mais recente (.xlsb ou .xlsx) contendo o fragmento no nome."""
    matches = [
        p for p in folder.iterdir()
        if p.suffix.lower() in (".xlsb", ".xlsx")
        and not p.name.startswith("~$")
        and fragment.lower() in p.name.lower()
    ]
    if not matches:
        raise FileNotFoundError(f"Arquivo contendo {fragment!r} não encontrado em {folder}")
    return max(matches, key=lambda p: p.stat().st_mtime)


def read_sheet(path: Path, sheet_name: str) -> list[dict[int, Any]]:
    rows: list[dict[int, Any]] = []
    with open_workbook(str(path)) as book:
        with book.get_sheet(sheet_name) as sheet:
            for row in sheet.rows():
                values = {cell.c: cell.v for cell in row if cell.v not in (None, "")}
                rows.append(values)
    return rows


def new_report(title: str, subtitle: str, end_col: int = 8) -> tuple[Workbook, Any]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    end_letter = get_column_letter(end_col)
    ws.merge_cells(f"A1:{end_letter}1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(f"A2:{end_letter}2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Aptos", size=10, italic=True, color="5B6573")
    ws.row_dimensions[2].height = 22
    return wb, ws


def style_table(ws, header_row: int, first_col: int, last_col: int, last_row: int, table_name: str) -> None:
    header = ws.cell(header_row, first_col).coordinate + ":" + ws.cell(header_row, last_col).coordinate
    for row in ws[header]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=BLUE)
            cell.font = Font(bold=True, color=WHITE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ref = f"{ws.cell(header_row, first_col).coordinate}:{ws.cell(last_row, last_col).coordinate}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
    ws.add_table(table)
    for row in ws.iter_rows(min_row=header_row + 1, max_row=last_row, min_col=first_col, max_col=last_col):
        for cell in row:
            cell.border = Border(bottom=THIN_GRAY)


def set_widths(ws, widths: dict[int, float]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def add_source_sheet(wb: Workbook, source: Path, details: Iterable[str]) -> None:
    ws = wb.create_sheet("Fonte")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Rastreabilidade"
    ws["A1"].font = Font(size=16, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A3"] = "Arquivo-fonte"
    ws["B3"] = str(source)
    ws["A4"] = "Última modificação"
    ws["B4"] = source.stat().st_mtime
    ws["B4"].number_format = "0"
    for idx, detail in enumerate(details, start=6):
        ws.cell(idx, 1, detail)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 95
    ws["B3"].alignment = Alignment(wrap_text=True)


def extract_air_freight_data(source: Path) -> list[dict]:
    rows = read_sheet(source, "Annual Result")
    tv_rate = rows[9]
    result_cols = list(range(29, 36))
    target = tv_rate.get(42)
    records = []
    for month, col in zip(MONTHS[:7], result_cols):
        result = tv_rate.get(col)
        if result is None:
            continue
        try:
            target_val = float(target) if target is not None else 0.0
            result_val = float(result)
            achievement = round(target_val / result_val, 4) if result_val > 0 else None
            records.append({"month": month, "year": "Y26", "target": target_val, "result": result_val, "achievement": achievement})
        except ValueError:
            pass
    return records


def build_air_freight(source: Path, output: Path) -> None:
    rows = read_sheet(source, "Annual Result")
    tv_rate = rows[9]  # Excel row 10: TV / %
    result_cols = list(range(29, 36))  # AD:AJ, Jan-Jul/2026, zero-based
    target = tv_rate.get(42)  # AQ

    wb, ws = new_report("KPI – Air Freight", "Produto TV | Annual Result | Jan–Jul 2026")
    ws.append([])
    ws.append(["Mês", "Target", "Result"])
    for month, col in zip(MONTHS[:7], result_cols):
        ws.append([month, target, tv_rate.get(col)])
    for cell in ws[4]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center")
    for row in range(5, 12):
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        ws.cell(row, 2).font = Font(color="666666")
        ws.cell(row, 3).fill = PatternFill("solid", fgColor=LIGHT_RED)
        ws.cell(row, 3).font = Font(color=RED, bold=True)
        for col in range(2, 4):
            ws.cell(row, col).number_format = "0.00%"
    style_table(ws, 4, 1, 3, 11, "AirFreightTV")
    set_widths(ws, {1: 13, 2: 15, 3: 15})
    add_source_sheet(wb, source, ["Aba: Annual Result", "Produto: TV", "Resultado 2026: AD:AJ", "Target: AQ"])
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(output)


def extract_incidental_cost_data(source: Path) -> list[dict]:
    rows = read_sheet(source, "Incidental Cost (MUSD)")
    incidental = rows[81]
    prod_amt = rows[94]
    periods = [(2025, month, col) for month, col in zip(MONTHS, range(20, 32))]
    periods += [(2026, month, col) for month, col in zip(MONTHS[:7], range(32, 39))]
    records = []
    for year, month, col in periods:
        inc = incidental.get(col)
        prod = prod_amt.get(col)
        if inc is None or prod is None:
            continue
        try:
            inc_val = float(inc)
            prod_val = float(prod)
            if prod_val > 0:
                result_val = inc_val / prod_val
                records.append({"month": month, "year": f"Y{str(year)[-2:]}", "target": 0.0, "result": result_val, "achievement": None})
        except ValueError:
            pass
    return records


def build_incidental_cost(source: Path, output: Path) -> None:
    rows = read_sheet(source, "Incidental Cost (MUSD)")
    incidental = rows[81]  # Excel row 82
    prod_amt = rows[94]  # Excel row 95
    periods = [(2025, month, col) for month, col in zip(MONTHS, range(20, 32))]
    periods += [(2026, month, col) for month, col in zip(MONTHS[:7], range(32, 39))]

    wb, ws = new_report("KPI – Incidental Cost", "Produto TV | Incidental Cost ÷ Prod. Amt. | 2025–Jul/2026")
    ws.append([])
    ws.append(["Ano", "Mês", "Incidental Cost", "Incidental Cost (MUSD)", "Prod. Amt. (MUSD)"])
    for year, month, col in periods:
        row = ws.max_row + 1
        ws.append([year, month, f'=IFERROR(D{row}/E{row},"")', incidental.get(col), prod_amt.get(col)])
    last_row = ws.max_row
    style_table(ws, 4, 1, 5, last_row, "IncidentalCostTV")
    for row in range(5, last_row + 1):
        ws.cell(row, 3).number_format = "0.0%"
        ws.cell(row, 3).fill = PatternFill("solid", fgColor=LIGHT_RED)
        ws.cell(row, 3).font = Font(color=RED, bold=True)
        ws.cell(row, 4).number_format = "0.00"
        ws.cell(row, 5).number_format = "0.00"
    set_widths(ws, {1: 10, 2: 12, 3: 20, 4: 24, 5: 22})
    add_source_sheet(wb, source, ["Aba: Incidental Cost (MUSD)", "Bloco de TV iniciado na linha 76", "Incidental Cost: linha 82", "Prod. Amt.: linha 95", "Percentual calculado por fórmula no arquivo de saída"])
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(output)


def extract_war_room_data(source: Path) -> list[dict]:
    rows = read_sheet(source, "Logistic")
    tv_row = rows[67]
    if str(tv_row.get(0, "")).strip().upper() != "TV":
        raise RuntimeError("A linha 68 da aba Logistic não está identificada como TV")
    records = []
    for idx, month in enumerate(MONTHS):
        res_col = 33 + idx
        res = tv_row.get(res_col)
        if res is not None:
            try:
                res_val = float(res)
                records.append({"month": month, "year": "Y25", "target": 0.0, "result": res_val, "achievement": None})
            except ValueError:
                pass
    for idx, month in enumerate(MONTHS):
        tgt_col = 59 + idx
        res_col = 72 + idx
        tgt = tv_row.get(tgt_col)
        res = tv_row.get(res_col)
        if res is not None:
            try:
                tgt_val = float(tgt) if tgt is not None else 0.0
                res_val = float(res)
                achievement = round(tgt_val / res_val, 4) if res_val > 0 else None
                records.append({"month": month, "year": "Y26", "target": tgt_val, "result": res_val, "achievement": achievement})
            except ValueError:
                pass
    return records


def build_war_room(source: Path, output: Path) -> None:
    rows = read_sheet(source, "Logistic")
    tv_row = rows[67]  # Excel row 68
    if str(tv_row.get(0, "")).strip().upper() != "TV":
        raise RuntimeError("A linha 68 da aba Logistic não está identificada como TV")

    wb, ws = new_report("KPI – War Room", "Logistic | Linha 68 | TV | Faixa AH:CG sem BP")
    headers = ["Produto", "Série"] + MONTHS + ["Accu"]
    ws.append([])
    ws.append(headers)
    series = [
        ("25Y Result", range(33, 46), LIGHT_BLUE, NAVY),  # AH:AT
        ("26Y Target", range(59, 72), LIGHT_GRAY, "666666"),  # BH:BT
        ("26Y Result", range(72, 85), LIGHT_RED, RED),  # BU:CG
    ]
    for label, cols, fill_color, font_color in series:
        out_row = ws.max_row + 1
        ws.append(["TV", label] + [tv_row.get(c) for c in cols])
        for col in range(2, 16):
            ws.cell(out_row, col).fill = PatternFill("solid", fgColor=fill_color)
            ws.cell(out_row, col).font = Font(color=font_color, bold=(label != "25Y Result"))
            if col >= 3:
                ws.cell(out_row, col).number_format = "0.00%"
    last_row = ws.max_row
    style_table(ws, 4, 1, 15, last_row, "WarRoomTV")
    set_widths(ws, {1: 12, 2: 17, **{col: 13 for col in range(3, 16)}})
    add_source_sheet(wb, source, ["Aba: Logistic", "Produto: TV", "Linha: 68", "Faixa consultada: AH:CG", "BP excluído conforme solicitação", "25Y Result: AH:AT", "26Y Target: BH:BT", "26Y Result: BU:CG"])
    wb.save(output)


def source_value(value: Any) -> float:
    """Converte números e textos monetários da planilha manual para float."""
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace("$", "").replace(" ", "").replace(".", "").replace(",", ".")
    return float(cleaned)


def style_indicator_rows(ws, first_row: int, last_row: int, status_col: int) -> None:
    for row in range(first_row, last_row + 1):
        estimated = ws.cell(row, status_col).value == "Estimado"
        fill = PatternFill("solid", fgColor="FFF2CC" if estimated else "E2F0D9")
        for col in range(1, status_col + 1):
            ws.cell(row, col).fill = fill
            ws.cell(row, col).border = Border(bottom=THIN_GRAY)


def style_summary(ws, header_row: int, last_col: int, last_row: int) -> None:
    for cell in ws[header_row][:last_col]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=header_row + 1, max_row=last_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.border = Border(bottom=THIN_GRAY)


def extract_task_cost_data(source: Path) -> list[dict]:
    src_wb = load_workbook(source, data_only=True, read_only=True)
    src = src_wb.active
    target25 = [src.cell(3, col).value for col in range(5, 17)]
    result25 = [src.cell(4, col).value for col in range(5, 17)]
    target26_raw = [src.cell(3, col).value for col in range(19, 31)]
    result26_raw = [src.cell(4, col).value for col in range(19, 31)]
    src_wb.close()

    known_target_months = sum(1 for value in target26_raw if value is not None)
    target_ratio = sum(target26_raw[:known_target_months]) / sum(target25[:known_target_months]) if known_target_months > 0 and sum(target25[:known_target_months]) > 0 else 1.0

    records = []
    for idx, month in enumerate(MONTHS):
        if result25[idx] is not None and target25[idx] is not None:
            try:
                tgt = float(target25[idx])
                res = float(result25[idx])
                achievement = round(res / tgt, 4) if tgt > 0 else None
                records.append({"month": month, "year": "Y25", "target": tgt, "result": res, "achievement": achievement})
            except ValueError:
                pass

    for idx, month in enumerate(MONTHS):
        target_status = "Informado" if target26_raw[idx] is not None else "Estimado"
        result_status = "Informado" if result26_raw[idx] is not None else "Estimado"
        
        tgt_val = target26_raw[idx] if target_status == "Informado" else round(target25[idx] * target_ratio) if target25[idx] is not None else 0.0
        if tgt_val is None:
            continue
            
        if result_status == "Informado":
            res_val = result26_raw[idx]
        else:
            if target25[idx] and target25[idx] > 0 and result25[idx] is not None:
                res_val = round(tgt_val * result25[idx] / target25[idx])
            else:
                res_val = 0.0
                
        if res_val is not None:
            try:
                tgt = float(tgt_val)
                res = float(res_val)
                achievement = round(res / tgt, 4) if tgt > 0 else None
                records.append({"month": month, "year": "Y26", "target": tgt, "result": res, "achievement": achievement})
            except ValueError:
                pass

    return records


def build_task_cost_reduction(source: Path, output: Path) -> None:
    src_wb = load_workbook(source, data_only=True, read_only=True)
    src = src_wb.active
    target25 = [src.cell(3, col).value for col in range(5, 17)]
    result25 = [src.cell(4, col).value for col in range(5, 17)]
    target26_raw = [src.cell(3, col).value for col in range(19, 31)]
    result26_raw = [src.cell(4, col).value for col in range(19, 31)]
    src_wb.close()

    known_target_months = sum(value is not None for value in target26_raw)
    target_ratio = sum(target26_raw[:known_target_months]) / sum(target25[:known_target_months])

    wb, ws = new_report(
        "Task Cost Reduction (Logistics) — Internal Operation",
        "Truck Head / Reachstacker / Chassis | Valores em KBRL | Campos ausentes de 2026 estimados e identificados",
        8,
    )
    ws.append([])
    ws.append(["Ano", "Mês nº", "Mês", "Target (KBRL)", "Result (KBRL)", "Atingimento", "Status Target", "Status Result"])
    for month_idx, month in enumerate(MONTHS):
        row = ws.max_row + 1
        ws.append([2025, month_idx + 1, month, target25[month_idx], result25[month_idx], f"=IFERROR(E{row}/D{row},0)", "Informado", "Informado"])
    for month_idx, month in enumerate(MONTHS):
        row = ws.max_row + 1
        target_status = "Informado" if target26_raw[month_idx] is not None else "Estimado"
        result_status = "Informado" if result26_raw[month_idx] is not None else "Estimado"
        target = target26_raw[month_idx] if target_status == "Informado" else round(target25[month_idx] * target_ratio)
        result = result26_raw[month_idx] if result_status == "Informado" else round(target * result25[month_idx] / target25[month_idx])
        ws.append([2026, month_idx + 1, month, target, result, f"=IFERROR(E{row}/D{row},0)", target_status, result_status])
    style_table(ws, 4, 1, 8, 28, "TaskCostReduction")
    style_indicator_rows(ws, 5, 28, 8)
    for row in range(5, 29):
        ws.cell(row, 4).number_format = "#,##0"
        ws.cell(row, 5).number_format = "#,##0"
        ws.cell(row, 6).number_format = "0.0%"
    ws.append([])
    ws.append([])
    ws.append(["Ano", "Target Total", "Result Total", "Atingimento", "Observação"])
    ws.append([2025, "=SUM(D5:D16)", "=SUM(E5:E16)", "=IFERROR(C32/B32,0)", "Dados informados no arquivo de origem"])
    ws.append([2026, "=SUM(D17:D28)", "=SUM(E17:E28)", "=IFERROR(C33/B33,0)", "Inclui estimativas identificadas para campos faltantes"])
    style_summary(ws, 31, 5, 33)
    for row in (32, 33):
        ws.cell(row, 2).number_format = "#,##0"
        ws.cell(row, 3).number_format = "#,##0"
        ws.cell(row, 4).number_format = "0.0%"
    set_widths(ws, {1: 15, 2: 12, 3: 10, 4: 17, 5: 17, 6: 15, 7: 16, 8: 16})
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(output)


def extract_demurrage_data(source: Path) -> list[dict]:
    src_wb = load_workbook(source, data_only=True, read_only=True)
    src = src_wb.active
    target25 = [src.cell(12, col).value or 0 for col in range(5, 17)]
    result25 = [src.cell(13, col).value or 0 for col in range(5, 17)]
    target26 = [src.cell(12, col).value or 0 for col in range(19, 31)]
    result26_raw = [src.cell(13, col).value for col in range(19, 31)]
    src_wb.close()

    records = []
    for idx, month in enumerate(MONTHS):
        res = result25[idx]
        if res is not None:
            try:
                tgt_val = float(target25[idx])
                res_val = float(res)
                achievement = round(tgt_val / res_val, 4) if res_val > 0 else None
                records.append({"month": month, "year": "Y25", "target": tgt_val, "result": res_val, "achievement": achievement})
            except ValueError:
                pass
    for idx, month in enumerate(MONTHS):
        res = result26_raw[idx] or 0
        if res is not None:
            try:
                tgt_val = float(target26[idx])
                res_val = float(res)
                achievement = round(tgt_val / res_val, 4) if res_val > 0 else None
                records.append({"month": month, "year": "Y26", "target": tgt_val, "result": res_val, "achievement": achievement})
            except ValueError:
                pass
    return records


def build_demurrage(source: Path, output: Path) -> None:
    src_wb = load_workbook(source, data_only=True, read_only=True)
    src = src_wb.active
    target25 = [src.cell(12, col).value or 0 for col in range(5, 17)]
    result25 = [src.cell(13, col).value or 0 for col in range(5, 17)]
    usd25 = [src.cell(14, col).value or 0 for col in range(5, 17)]
    target26 = [src.cell(12, col).value or 0 for col in range(19, 31)]
    result26_raw = [src.cell(13, col).value for col in range(19, 31)]
    usd26_raw = [src.cell(14, col).value for col in range(19, 31)]
    src_wb.close()

    wb, ws = new_report(
        "KPI — Demurrage Cost (TV)",
        "Target e Result em quantidade de contêineres; gasto em USD. Lacunas de 2026 preenchidas com zero esperado e identificadas como estimativa.",
        9,
    )
    ws.append([])
    ws.append(["Ano", "Mês nº", "Mês", "Target (CTNR)", "Result (CTNR)", "Valor gasto (USD)", "Status Target", "Status Result", "Status USD"])
    for month_idx, month in enumerate(MONTHS):
        ws.append([2025, month_idx + 1, month, target25[month_idx], result25[month_idx], usd25[month_idx], "Informado", "Informado", "Informado"])
    for month_idx, month in enumerate(MONTHS):
        result_status = "Informado" if result26_raw[month_idx] is not None else "Estimado"
        usd_status = "Informado" if result26_raw[month_idx] is not None else "Estimado"
        ws.append([2026, month_idx + 1, month, target26[month_idx], result26_raw[month_idx] or 0, usd26_raw[month_idx] or 0, "Informado", result_status, usd_status])
    style_table(ws, 4, 1, 9, 28, "DemurrageCostTV")
    style_indicator_rows(ws, 5, 28, 9)
    for row in range(5, 29):
        ws.cell(row, 4).number_format = "#,##0"
        ws.cell(row, 5).number_format = "#,##0"
        ws.cell(row, 6).number_format = '"$"#,##0.00'
    ws.append([])
    ws.append([])
    ws.append(["Ano", "Target Total", "Result Total", "Valor gasto total (USD)"])
    ws.append([2025, "=SUM(D5:D16)", "=SUM(E5:E16)", "=SUM(F5:F16)"])
    ws.append([2026, "=SUM(D17:D28)", "=SUM(E17:E28)", "=SUM(F17:F28)"])
    style_summary(ws, 31, 4, 33)
    for row in (32, 33):
        ws.cell(row, 4).number_format = '"$"#,##0.00'
    set_widths(ws, {1: 15, 2: 12, 3: 10, 4: 16, 5: 16, 6: 19, 7: 16, 8: 16, 9: 16})
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(output)


def extract_resin_data(source: Path) -> list[dict]:
    src_wb = load_workbook(source, data_only=True, read_only=True)
    src = src_wb.active
    
    records = []
    
    # 2025: Coluna F (6) até Q (17) -> jan a dez
    for i, month in enumerate(MONTHS):
        col = 6 + i
        target_val = src.cell(row=23, column=col).value
        cost1_val = src.cell(row=24, column=col).value
        cost2_val = src.cell(row=25, column=col).value
        
        try:
            t = float(target_val) if target_val is not None else 0.0
            c1 = float(cost1_val) if cost1_val is not None else 0.0
            c2 = float(cost2_val) if cost2_val is not None else 0.0
            
            # Aplica a fórmula segura: Linha 23 - Linha 24 - Linha 25
            r = round(t - c1 - c2, 2)
            
            if t != 0.0 or r != 0.0:
                records.append({
                    "month": month,
                    "year": "Y25",
                    "target": t,
                    "result": r,
                    "achievement": None
                })
        except (ValueError, TypeError):
            pass
            
    # 2026: Coluna T (20) até AE (31) -> jan a dez
    for i, month in enumerate(MONTHS):
        col = 20 + i
        target_val = src.cell(row=23, column=col).value
        cost1_val = src.cell(row=24, column=col).value
        cost2_val = src.cell(row=25, column=col).value
        
        try:
            t = float(target_val) if target_val is not None else 0.0
            c1 = float(cost1_val) if cost1_val is not None else 0.0
            c2 = float(cost2_val) if cost2_val is not None else 0.0
            
            # Aplica a fórmula segura: Linha 23 - Linha 24 - Linha 25
            r = round(t - c1 - c2, 2)
            
            if t != 0.0 or r != 0.0:
                records.append({
                    "month": month,
                    "year": "Y26",
                    "target": t,
                    "result": r,
                    "achievement": None
                })
        except (ValueError, TypeError):
            pass

    src_wb.close()
    return records


def build_resin_consolidation(source: Path, output: Path) -> None:
    src_wb = load_workbook(source, data_only=True, read_only=True)
    src = src_wb.active
    data25 = [[source_value(src.cell(row, col).value) for col in range(6, 18)] for row in range(21, 27)]
    data26 = [[src.cell(row, col).value for col in range(20, 32)] for row in range(21, 27)]
    src_wb.close()

    known_months = sum(value is not None for value in data26[0])
    ratio40 = sum(data26[0][:known_months]) / sum(data25[0][:known_months])
    ratio20 = sum(data26[1][:known_months]) / sum(data25[1][:known_months])
    unit_cost = sum(data26[3][:known_months]) / sum(data26[0][:known_months])

    wb, ws = new_report(
        "Logistics Cost Resin Consolidation",
        "Valores financeiros em KUSD. Campos ausentes de 2026 seguem a sazonalidade de 2025 ajustada pelo desempenho Jan–Abr/2026.",
        10,
    )
    ws.append([])
    ws.append(["Ano", "Mês nº", "Mês", "CTNs 40 ft", "CTNs Saving 20 ft", "Saving Valor (KUSD)", "Consolidation Costs (KUSD)", "BR Tax 34,39% (KUSD)", "Saving líquido (KUSD)", "Status"])
    for month_idx, month in enumerate(MONTHS):
        ws.append([2025, month_idx + 1, month, data25[0][month_idx], data25[1][month_idx], data25[2][month_idx], data25[3][month_idx], data25[4][month_idx], data25[5][month_idx], "Informado"])
    for month_idx, month in enumerate(MONTHS):
        row = ws.max_row + 1
        informed = data26[0][month_idx] is not None
        qty40 = data26[0][month_idx] if informed else round(data25[0][month_idx] * ratio40)
        qty20 = data26[1][month_idx] if informed else round(data25[1][month_idx] * ratio20)
        if informed:
            ws.append([2026, month_idx + 1, month, qty40, qty20, data26[2][month_idx], data26[3][month_idx], data26[4][month_idx], data26[5][month_idx], "Informado"])
        else:
            ws.append([2026, month_idx + 1, month, qty40, qty20, f"=ROUND(E{row}*3.958,2)", round(qty40 * unit_cost, 2), f"=ROUND(G{row}*34.39%,2)", f"=ROUND(F{row}-G{row}-H{row},2)", "Estimado"])
    style_table(ws, 4, 1, 10, 28, "ResinConsolidation")
    style_indicator_rows(ws, 5, 28, 10)
    for row in range(5, 29):
        ws.cell(row, 4).number_format = "#,##0"
        ws.cell(row, 5).number_format = "#,##0"
        for col in range(6, 10):
            ws.cell(row, col).number_format = '"$"#,##0.00'
    ws.append([])
    ws.append([])
    ws.append(["Ano", "CTNs 40 ft", "CTNs Saving 20 ft", "Saving Valor", "Costs", "BR Tax", "Saving líquido"])
    ws.append([2025, "=SUM(D5:D16)", "=SUM(E5:E16)", "=SUM(F5:F16)", "=SUM(G5:G16)", "=SUM(H5:H16)", "=SUM(I5:I16)"])
    ws.append([2026, "=SUM(D17:D28)", "=SUM(E17:E28)", "=SUM(F17:F28)", "=SUM(G17:G28)", "=SUM(H17:H28)", "=SUM(I17:I28)"])
    style_summary(ws, 31, 7, 33)
    for row in (32, 33):
        ws.cell(row, 2).number_format = "#,##0"
        ws.cell(row, 3).number_format = "#,##0"
        for col in range(4, 8):
            ws.cell(row, col).number_format = '"$"#,##0.00'
    set_widths(ws, {1: 13, 2: 12, 3: 10, 4: 16, 5: 19, 6: 21, 7: 23, 8: 22, 9: 21, 10: 14})
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(output)


def verify_workbook(path: Path, expected_sheet: str = "Dados") -> None:
    wb = load_workbook(path, data_only=False, read_only=False)
    if expected_sheet not in wb.sheetnames:
        raise RuntimeError(f"Aba {expected_sheet!r} ausente em {path.name}")
    ws = wb[expected_sheet]
    if ws.max_row < 5 or ws.max_column < 4:
        raise RuntimeError(f"Saída aparentemente vazia: {path.name}")
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and any(err in cell.value for err in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?")):
                raise RuntimeError(f"Erro de fórmula em {path.name} {cell.coordinate}: {cell.value}")
    wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Extrai os KPIs e recria as seis planilhas oficiais em saidas_kpi.")
    parser.add_argument("--origem", type=Path, default=DEFAULT_SOURCE, help="Pasta que contém os arquivos XLSB")
    parser.add_argument("--saida", type=Path, default=DEFAULT_OUTPUT, help="Pasta de destino das planilhas")
    parser.add_argument("--somente-war-room", action="store_true", help="Gera somente a planilha War Room")
    args = parser.parse_args()
    args.saida.mkdir(parents=True, exist_ok=True)

    war_room = find_file(args.origem, "War Room")

    if args.somente_war_room:
        output = args.saida / "KPI_War_Room_TV.xlsx"
        build_war_room(war_room, output)
        verify_workbook(output)
        print(f"OK: {output}")
        return

    freight = find_file(args.origem, "Freight Air (Monthly)")
    incidental = find_file(args.origem, "Incidental Cost_Total_v0")
    three_indicators = find_xlsx(args.origem, "3-indicadores")

    outputs = [
        args.saida / "KPI_Air_Freight_TV_2026.xlsx",
        args.saida / "KPI_Incidental_Cost_TV.xlsx",
        args.saida / "KPI_War_Room_TV.xlsx",
        args.saida / "KPI_Task_Cost_Reduction_Logistics.xlsx",
        args.saida / "KPI_Demurrage_Cost_TV.xlsx",
        args.saida / "KPI_Logistics_Cost_Resin_Consolidation.xlsx",
    ]
    build_air_freight(freight, outputs[0])
    build_incidental_cost(incidental, outputs[1])
    build_war_room(war_room, outputs[2])
    build_task_cost_reduction(three_indicators, outputs[3])
    build_demurrage(three_indicators, outputs[4])
    build_resin_consolidation(three_indicators, outputs[5])
    for output in outputs:
        verify_workbook(output)
        print(f"OK: {output}")


if __name__ == "__main__":
    main()
