from pathlib import Path

from openpyxl import load_workbook

from rpa_email.extractors.common import MONTHS, monthly_record


KPI_KEY = "demurrage"
DISPLAY_NAME = "Demurrage Cost"


def extract(source: Path) -> list[dict]:
    """Extrai Demurrage Cost do arquivo 3-indicadores."""

    workbook = load_workbook(source, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        target_2025 = [sheet.cell(12, column).value for column in range(5, 17)]
        result_2025 = [sheet.cell(13, column).value for column in range(5, 17)]
        target_2026 = [sheet.cell(12, column).value for column in range(19, 31)]
        result_2026 = [sheet.cell(13, column).value for column in range(19, 31)]
    finally:
        workbook.close()

    records = []
    for year, targets, results in (
        ("Y25", target_2025, result_2025),
        ("Y26", target_2026, result_2026),
    ):
        for month, target, result in zip(MONTHS, targets, results):
            if target is None or result is None:
                continue
            records.append(monthly_record(month, year, target, result, lower_is_better=True))
    return records
