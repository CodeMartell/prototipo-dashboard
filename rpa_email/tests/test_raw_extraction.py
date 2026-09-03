import openpyxl
import pytest

from rpa_email.extractors import RawReportExtractor
from rpa_email.services.api_client import IngestionError, build_payload


def _save_raw_workbooks(folder):
    war_room = openpyxl.Workbook()
    war_sheet = war_room.active
    war_sheet.title = "Logistic"
    war_sheet.cell(68, 1, "TV")
    war_sheet.cell(68, 21, 0.06)  # U, Jan/Y24
    war_sheet.cell(68, 34, 0.05)  # AH, Jan/Y25
    war_sheet.cell(68, 60, 0.04)  # BH, target Jan/Y26
    war_sheet.cell(68, 73, 0.03)  # BU, result Jan/Y26
    war_sheet.cell(68, 83, 0)     # CE, Nov/Y26 ainda sem dado
    war_room.save(folder / "War Room.xlsx")
    war_room.close()

    freight = openpyxl.Workbook()
    freight_sheet = freight.active
    freight_sheet.title = "Annual Result"
    freight_sheet.cell(10, 30, 0.005)  # AD, Jan/Y26
    freight_sheet.cell(10, 43, 0.0022)  # AQ, target
    freight.save(folder / "Freight Air.xlsx")
    freight.close()

    cost_vs_prod = openpyxl.Workbook()
    cost_sheet = cost_vs_prod.active
    cost_sheet.title = "Incidental Cost (MUSD)"
    cost_sheet.cell(6, 21, "2025-01")
    cost_sheet.cell(82, 21, 2)   # U, Jan/Y25 — custo
    cost_sheet.cell(95, 21, 10)  # U, Jan/Y25 — produção
    cost_sheet.cell(96, 21, 0.2)  # U, Jan/Y25 — percentual publicado
    cost_sheet.cell(6, 33, "2026-01")
    cost_sheet.cell(82, 33, 3)   # AG, Jan/Y26 — custo
    cost_sheet.cell(95, 33, 12)  # AG, Jan/Y26 — produção
    cost_sheet.cell(96, 33, 0.25)  # AG, Jan/Y26 — percentual publicado
    cost_sheet.cell(6, 34, "2026-02")
    cost_sheet.cell(82, 34, 4)   # fórmula/projeção futura sem produção publicada
    cost_vs_prod.save(folder / "_26.07 Incidental Cost_Total_v0.xlsx")
    cost_vs_prod.close()

    indicators = openpyxl.Workbook()
    sheet = indicators.active
    sheet.cell(3, 5, 100)
    sheet.cell(4, 5, 120)
    sheet.cell(3, 19, 110)
    sheet.cell(4, 19, 121)
    sheet.cell(12, 5, 0)
    sheet.cell(13, 5, 0)
    sheet.cell(12, 19, 0)
    sheet.cell(13, 19, 2)
    sheet.cell(23, 6, 10)
    sheet.cell(24, 6, 2)
    sheet.cell(25, 6, 1)
    sheet.cell(23, 20, 12)
    sheet.cell(24, 20, 3)
    sheet.cell(25, 20, 1)
    indicators.save(folder / "3-indicadores.xlsx")
    indicators.close()


def test_raw_xlsx_files_are_routed_to_domain_extractors(tmp_path):
    _save_raw_workbooks(tmp_path)

    result = RawReportExtractor().extract(tmp_path)

    assert result.errors == []
    assert len(result.logistic_cost) == 3
    assert len(result.air_freight) == 1
    assert len(result.total_cost) == 2
    assert len(result.demurrage) == 2
    assert len(result.incidental_cost) == 2
    assert len(result.logistics_vs_prod) == 2
    assert result.total_cost[0].achievement == 1.2
    assert result.demurrage[0].achievement == 1.0
    assert result.logistics_vs_prod[0].ratio == 0.2
    assert result.logistics_vs_prod[1].ratio == 0.25
    assert result.replace_kpis == {"logistic_cost"}
    assert [(row.year, row.month) for row in result.logistic_cost] == [
        ("Y24", "Jan"),
        ("Y25", "Jan"),
        ("Y26", "Jan"),
    ]


def test_raw_extractor_does_not_estimate_missing_months(tmp_path):
    _save_raw_workbooks(tmp_path)

    result = RawReportExtractor().extract(tmp_path)

    assert {row.month for row in result.total_cost} == {"Jan"}
    assert {row.month for row in result.demurrage} == {"Jan"}


def test_raw_extractor_accepts_currency_strings(tmp_path):
    _save_raw_workbooks(tmp_path)
    path = tmp_path / "3-indicadores.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    sheet.cell(23, 6, "$4,0 ")
    sheet.cell(24, 6, "R$ 1,25")
    sheet.cell(25, 6, "$0.50")
    workbook.save(path)
    workbook.close()

    result = RawReportExtractor().extract(tmp_path)

    assert result.errors == []
    assert result.incidental_cost[0].target == 4.0
    assert result.incidental_cost[0].result == 2.25


def test_incidental_cost_rejects_reported_ratio_mismatch(tmp_path):
    _save_raw_workbooks(tmp_path)
    path = tmp_path / "_26.07 Incidental Cost_Total_v0.xlsx"
    workbook = openpyxl.load_workbook(path)
    workbook["Incidental Cost (MUSD)"].cell(96, 33, 0.20)
    workbook.save(path)
    workbook.close()

    result = RawReportExtractor().extract(tmp_path)

    assert result.logistics_vs_prod == []
    assert any("percentual divergente em Jan/Y26" in error for error in result.errors)


def test_incidental_cost_rejects_zero_production_amount(tmp_path):
    _save_raw_workbooks(tmp_path)
    path = tmp_path / "_26.07 Incidental Cost_Total_v0.xlsx"
    workbook = openpyxl.load_workbook(path)
    workbook["Incidental Cost (MUSD)"].cell(95, 33, 0)
    workbook.save(path)
    workbook.close()

    result = RawReportExtractor().extract(tmp_path)

    assert result.logistics_vs_prod == []
    assert any("custo/produção inválido em Jan/Y26" in error for error in result.errors)


def test_missing_raw_source_rejects_entire_batch(tmp_path):
    _save_raw_workbooks(tmp_path)
    (tmp_path / "_26.07 Incidental Cost_Total_v0.xlsx").unlink()

    result = RawReportExtractor().extract(tmp_path)

    assert result.logistic_cost == []
    assert result.logistics_vs_prod == []
    assert result.errors == [
        "Relatórios brutos obrigatórios ausentes: Incidental Cost_Total"
    ]


def test_invalid_raw_value_blocks_partial_payload(tmp_path):
    _save_raw_workbooks(tmp_path)
    path = tmp_path / "3-indicadores.xlsx"
    workbook = openpyxl.load_workbook(path)
    workbook.active.cell(4, 5, -1)
    workbook.save(path)
    workbook.close()

    result = RawReportExtractor().extract(tmp_path)

    assert any("total_cost" in error for error in result.errors)
    with pytest.raises(IngestionError, match="Falha na extracao"):
        build_payload(result, "message-id", "subject", "sender@example.com")
