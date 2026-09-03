from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def parse_number(value: object) -> float:
    """Converte números e textos monetários em float.

    Aceita tanto a notação brasileira (``R$ 1.234,56``) quanto a americana
    (``$1,234.56``), além de espaços introduzidos por fórmulas/exports.
    """

    if isinstance(value, bool) or value is None:
        raise ValueError(f"valor numérico inválido: {value!r}")
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        raw = str(value).replace("\u00a0", " ").strip()
        negative = raw.startswith("(") and raw.endswith(")")
        if negative:
            raw = raw[1:-1].strip()
        raw = re.sub(r"(?i)\b(?:BRL|USD)\b", "", raw)
        raw = raw.replace("R$", "").replace("US$", "").replace("$", "").replace(" ", "")
        if not raw or not re.fullmatch(r"[+-]?[0-9][0-9.,]*", raw):
            raise ValueError(f"valor numérico inválido: {value!r}")

        if "," in raw and "." in raw:
            decimal_separator = "," if raw.rfind(",") > raw.rfind(".") else "."
            thousands_separator = "." if decimal_separator == "," else ","
            raw = raw.replace(thousands_separator, "").replace(decimal_separator, ".")
        elif "," in raw:
            parts = raw.split(",")
            raw = "".join(parts[:-1]) + "." + parts[-1]
        elif raw.count(".") > 1:
            parts = raw.split(".")
            raw = "".join(parts[:-1]) + "." + parts[-1]

        number = float(raw)
        if negative:
            number = -number

    if not math.isfinite(number):
        raise ValueError(f"valor numérico inválido: {value!r}")
    return number


def read_sheet(path: Path, sheet_name: str) -> list[dict[int, Any]]:
    """Lê uma aba preservando índices de coluna baseados em zero.

    Relatórios brutos podem chegar em XLSB ou XLSX. A implementação antiga
    anunciava suporte aos dois formatos, mas tentava abrir ambos com pyxlsb.
    """

    if path.suffix.lower() == ".xlsb":
        from pyxlsb import open_workbook

        rows: list[dict[int, Any]] = []
        with open_workbook(str(path)) as book:
            with book.get_sheet(sheet_name) as sheet:
                for row in sheet.rows():
                    rows.append({cell.c: cell.v for cell in row if cell.v not in (None, "")})
        return rows

    if path.suffix.lower() == ".xlsx":
        workbook = load_workbook(path, data_only=True, read_only=True)
        try:
            sheet = workbook[sheet_name]
            return [
                {index: value for index, value in enumerate(row) if value not in (None, "")}
                for row in sheet.iter_rows(values_only=True)
            ]
        finally:
            workbook.close()

    raise ValueError(f"Formato não suportado: {path.suffix}")


def monthly_record(month: str, year: str, target: object, result: object, *, lower_is_better: bool) -> dict:
    target_value = parse_number(target) if target is not None else 0.0
    result_value = parse_number(result)
    if not all(math.isfinite(value) and value >= 0 for value in (target_value, result_value)):
        raise ValueError(f"target/result inválido em {month}/{year}")
    if lower_is_better:
        achievement = (
            1.0 if target_value == result_value == 0
            else round(target_value / result_value, 4) if result_value > 0
            else None
        )
    else:
        achievement = (
            1.0 if target_value == result_value == 0
            else round(result_value / target_value, 4) if target_value > 0
            else None
        )
    return {
        "month": month,
        "year": year,
        "target": target_value,
        "result": result_value,
        "achievement": achievement,
    }
